import cv2
import math
import numpy as np
import os
import time
import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

F1=90
F2=90
tmp1=0
tmp2 = 0
tmp3=0
mapd=4.6
resolution_w=1280#320#1280
ftbp=2
map_w=round(resolution_w*3/2)
wdth_apprx=round(resolution_w*0.5625)
map_h=round(map_w/mapd)#
#tmul = np.float32([[1, 0, 16], [0, 1, -4]])
#tmfl = np.float32([[1, 0, 12], [0, 1, 4]])  # tmul[:, :-1], ([0],[13])
#320
#tmul = np.float32([[1, 0, 16], [0, 1, 20]])
#tmfl = np.float32([[1, 0, 12], [0, 1, 28]])  # tmul[:, :-1], ([0],[13])
#1280
tmul = np.float32([[1, 0, round(16/(1280/resolution_w))], [0, 1, round(20/(1280/resolution_w))]])
tmfl = np.float32([[1, 0, round(-60/(1280/resolution_w))], [0, 1, round(20/(1280/resolution_w))]])  # tmul[:, :-1], ([0],[13])
#crop nozzle from LK
Nsp=[round(wdth_apprx/2) - round(0.208*wdth_apprx),round(wdth_apprx/2) + round(0.208*wdth_apprx)] #Nsp=[round(wdth_apprx/2) - round(0.208*wdth_apprx),round(wdth_apprx/2) + round(0.208*wdth_apprx)]
Nep=[round(resolution_w/2) - round(0.117*resolution_w),round(resolution_w/2) + round(0.117*resolution_w)] #Nep=round(resolution_w/2) - round(0.117*resolution_w),round(resolution_w/2) + round(0.117*resolution_w)
Zerofill=np.zeros(round(0.208*wdth_apprx)+round(0.117*resolution_w)) #Zerofill=np.zeros(round(0.208*wdth_apprx)+round(0.117*resolution_w))

CDlineTh=1
IAlineTh=1

ref_wdth=0.428#428#0.285
#th = 420  # 425
th=365#420#45#420
thdet=121#121#15#121

obs_s=969
obs_f=1249

thmp=95#156#16#156
HTr=[0,120]#[5,14]#[95,120]
fontSc=0.5*resolution_w/320
thickn=int(1*resolution_w/320)
LK=[]
avheight, prvpp=[], []

excel='blue.xlsx'
SourceF='./blue'
ResF='./bluet_1280_2'#'./CV_video_Sept_1280px'#'./CV_video_RTv2_7'#Thickness16'
if not os.path.exists(ResF):
    os.makedirs(ResF)
else:
    print('Folder already exists. Delete to run again..')
    exit()

l1,l2 =[],[]
df = pd.DataFrame({'Print': l1, 'Average width': l2})

File_name=os.listdir('./'+SourceF)
'#####################'
def pre_render_static_text(fontSc, thickn):
    static_text_img = np.zeros((map_h, int(resolution_w*3/2), 3), dtype=np.uint8)  # Adjust size as needed
    #cv2.putText(static_text_img, 'Undistrupted line: ', (round(0.05 * map_w), round(0.86*map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc, (255, 255, 255), thickn, cv2.LINE_AA)
    #cv2.putText(static_text_img, 'Final path: ', (round(0.48*map_w), round(0.05 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc, (255, 255, 255), thickn, cv2.LINE_AA)
    return static_text_img
'######################'
# Overlay static text and draw dynamic parts
def draw_dynamic_texts(dd_sk, static_text_img, dr, drift, fontSc, thickn):
    # Overlay static text
    dd_sk = cv2.addWeighted(dd_sk, 1.0, static_text_img, 1.0, 0)

    # Draw dynamic parts
    cv2.putText(dd_sk, f'{dr} mm', (round(0.2 * map_w), round(0.89*map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc, (0, 255, 0), thickn, cv2.LINE_AA)
    cv2.putText(dd_sk, f'{drift} mm', (round(0.57*map_w), round(0.11 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc, (0, 255, 0), thickn, cv2.LINE_AA)

    return dd_sk

'###############ab##################'
def image_resize(image, width=None, height=None, inter=cv2.INTER_AREA): 
    # Grab the image size
    (h, w) = image.shape[:2]

    # If both width and height are None, return the original image
    if width is None and height is None:
        return image

    # Calculate the aspect ratio
    aspect_ratio = float(w) / h

    # Calculate new dimensions based on width or height
    if width is None:
        new_width = round(height * aspect_ratio)
        new_height = height
    elif height is None:
        new_width = width
        new_height = round(width / aspect_ratio)
    else:
        new_width = width
        new_height = height

    # Resize the image
    resized = cv2.resize(image, (new_width, new_height), interpolation=inter)

    return resized

'#################################'
def dscale(image, Blower, Bupper): 
    # Convert image to HSV color space
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Threshold the image to obtain a mask
    mask = cv2.inRange(hsv, Blower, Bupper)

    # Apply median blur to the mask
    mask_blurred = cv2.medianBlur(mask, 9)

    # Detect circles using Hough transform
    circles = cv2.HoughCircles(mask_blurred, cv2.HOUGH_GRADIENT, 8, 1, param1=100, param2=50)

    # If a circle is found, convert its parameters to integers and draw it on the image
    if circles is not None:
        circle = circles[0, 0]
        x, y, r = map(int, circle)
        cv2.circle(image, (x, y), r, (0, 255, 0), 4)

        # Convert circle parameters to integers
        circles = circle.astype(int)

    return circles, image, mask_blurred

'#################################'
def Fwidth(image, circle, cnt, newpath):  
    x, y, r = circle[:]  # Simplify tuple unpacking
    # Extract the region of interest from the image
    img = image[y - round(0.21*wdth_apprx):y + round(0.21*wdth_apprx), x - round(0.2578*resolution_w):x - round(0.0625*resolution_w)]
    # Optionally, save the extracted region as an image
    if cnt == th:
        cv2.imwrite(f"{newpath}/Mask_example_{tname[0]}.jpg", img)
    # Count the non-zero pixels along the columns
    how_many_c = np.count_nonzero(img, axis=0)
    # Calculate the average width
    real_hmc = how_many_c[how_many_c != 0]
    avwidthc = np.average(real_hmc / (2 * r)) * ref_wdth
    return round(avwidthc, 3)  # Round only when returning the result

'#################################'
def refpath(Limg, cnt, prvpp, color=(255, 0, 0), thickness=2):

    #print(cnt)
    positions = np.nonzero(Limg.T)
    if cnt<th:
        if not positions[1].any():
            ref = np.zeros_like(Limg)
        else:
            avheight.append(round((min(positions[2])+max(positions[2]))/2))
            p1=(min(positions[1]),round(np.average(avheight)*1.03))
            p2=(max(positions[1]),round(np.average(avheight)*1.03))
            scl = (p2[0] - p1[0]) / 18
            ref = cv2.line(np.zeros_like(Limg), p1, p2, color, thickness)
            prvpp=[p1, p2, scl]
    else:
        p3 = tuple(np.subtract(prvpp[1], (0, round(1.7 * prvpp[2]))))
        p4 = tuple(np.subtract(p3, (round(3 * prvpp[2]), 0)))
        p5 = tuple(map(sum, zip(p4, (0, round(0.5 * prvpp[2])))))
        p6 = tuple(np.subtract(p5, (round(3 * prvpp[2]), 0)))
        p7 = tuple(map(sum, zip(p6, (0, round(0.5 * prvpp[2])))))
        p8 = tuple(np.subtract(p7, (round(3 * prvpp[2]), 0)))

        ref = cv2.line(np.zeros_like(Limg), prvpp[0], prvpp[1], color, thickness)
        ref = cv2.line(ref, prvpp[1], p3, color, thickness)
        #ref = cv2.addWeighted(ref, alpha, msk, 1 - alpha, 0)
        ref = cv2.line(ref, p3, p4, color, thickness)
        #ref = cv2.addWeighted(ref, alpha, msk, 1 - alpha, 0)
        ref = cv2.line(ref, p4, p5, color, thickness)
        #ref = cv2.addWeighted(ref, alpha, msk, 1 - alpha, 0)
        ref = cv2.line(ref, p5, p6, color,thickness) #-168
        #ref = cv2.addWeighted(ref, alpha, msk, 1 - alpha, 0)
        ref = cv2.line(ref, p6, p7, color,thickness) #-28
        #ref = cv2.addWeighted(ref, alpha, msk, 1 - alpha, 0)
        ref = cv2.line(ref, p7, p8, color,thickness) #-168
        #ref = cv2.addWeighted(ref, alpha, msk, 1 - alpha, 0)

    return ref, prvpp
'#################################'
def crtmap2(comparison, segment):
    merge_skel = []
    cmp=comparison
    #for i, cmp in enumerate(comparison):
    if len(cmp.shape)>2:
        img = cv2.cvtColor(cmp, cv2.COLOR_BGR2GRAY)
    else:
        img=cmp
    _, gray = cv2.threshold(img, 1, 255, cv2.THRESH_BINARY)
    img = cv2.GaussianBlur(gray, (5, 5), 0)
    line = cv2.ximgproc.thinning(img)
    if segment == 0:
        line = cv2.merge((np.zeros_like(line), np.zeros_like(line), line))
    elif segment == 1:
        # img=255-img
        line = cv2.merge((np.zeros_like(line), line, np.zeros_like(line)))

    #merge_skel.append(mdskl)
    return line#merge_skel
'####################'
def averages(p):
    unique_positions, indices = np.unique(p[0], return_inverse=True)
    sums = np.zeros_like(unique_positions, dtype=np.float64)
    counts = np.zeros_like(unique_positions, dtype=np.int32)
    np.add.at(sums, indices, p[1])
    np.add.at(counts, indices, 1)
    averages = sums / counts
    return np.column_stack((unique_positions, averages)).reshape(-1, 2)

'#################################'
def Ferror(reference, image, circle, ref_wdth):
    drift = []

    # Segmentation
    hl1_r, hl1_i = reference[round(0.39*map_h):, :round(0.84*map_w)], image[round(0.39*map_h):, :round(0.84*map_w)]
    hl234_r, hl234_i = reference[:round(0.39*map_h), :round(0.84*map_w)], image[:round(0.39*map_h), :round(0.84*map_w)]
    vl1_r, vl1_i = reference[:, round(0.84*map_w):], image[:, round(0.84*map_w):]
    #if cnt>=1100 and cnt<=1105:
    #    print('h')
    Segm = [(hl1_r, hl1_i), (hl234_r, hl234_i), (vl1_r, vl1_i)]
    for i, (seg_r, seg_i) in enumerate(Segm):
        # Non-zero positions
        positions_ref = np.nonzero(seg_r.T) if i < 2 else np.nonzero(seg_r)
        positions_img = np.nonzero(seg_i.T) if i < 2 else np.nonzero(seg_i)

        if not positions_img[0].any() or not positions_ref[0].any():
            continue

        # Calculate averages
        res_r = averages(positions_ref)
        res_i = averages(positions_img)

        # Get the minimum size
        ms = min(res_i.shape[0], res_r.shape[0])

        if res_i[0, 0] > res_r[0, 0]:
            startP = np.searchsorted(res_r[:, 0], res_i[0, 0])
            endP = min(ms + startP, res_r.shape[0])
            if endP > res_i.shape[0] and startP <res_i.shape[0]:
                #if cnt>=938 and cnt<=944:
                #    print('opa')
                drift_segment = np.mean(res_r[startP:res_i.shape[0], 1] - res_i[:ms-startP, 1])
            elif startP > ms:
                #print(cnt)
                #if cnt>=938 and cnt<=944:
                #    print('opa')
                drift_segment = np.mean(res_r[startP:endP, 1] - res_i[:, 1])
            else:
                drift_segment = np.mean(res_r[startP:endP, 1] - res_i[:ms-startP, 1])

        else:
            startP = np.searchsorted(res_i[:, 0], res_r[0, 0])
            endP = min(ms + startP, res_i.shape[0])
            if endP > res_i.shape[0] or startP==endP:
                continue
            else:
                drift_segment = np.mean(res_i[startP:endP, 1] - res_r[:endP-startP, 1])

        drift.append(drift_segment)

    r = circle[2]
    #print("Drift values:", drift)
    if not drift:
        error = None
    else:
        drift=[x for x in drift if str(x) != 'nan']
        error = round((sum(map(abs, drift)) / len(drift) / (2 * r)) * ref_wdth, 3)
    #print("Error:", error)
    return error
'##########################'
def mapping(oldfrm,newfrm, cnt):
    #newfrm[0:4, 0:] = 0
    #oldfrm[0:4, 0:] = 0
    newfrm=newfrm[:,round(0.1*resolution_w):round(1*resolution_w/2)]
    oldfrm=oldfrm[:,round(0.1*resolution_w):round(1*resolution_w/2)]
    listX, listY=[], []
    # params for ShiTomasi corner detection
    feature_params = dict(maxCorners=500, qualityLevel=0.01, minDistance=10, blockSize=10)
    # Parameters for lucas kanade optical flow
    lk_params = dict(winSize=(120, 120), maxLevel=2, criteria=(cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, 10, 0.03))
    #Random colors propably for the lines
    color = np.random.randint(0, 255, (1000, 3))

    if len(oldfrm.shape) > 2:
        GRold = cv2.cvtColor(oldfrm, cv2.COLOR_BGR2GRAY)
    else:
        GRold = oldfrm
    p0 = cv2.goodFeaturesToTrack(GRold, mask=None, **feature_params)

    mask = np.zeros_like(GRold)
    mask = cv2.merge((mask, mask, mask))
    frame = cv2.merge((newfrm, newfrm, newfrm))
    if len(newfrm.shape) > 2:
        GRnew = cv2.cvtColor(newfrm, cv2.COLOR_BGR2GRAY)
    else:
        GRnew = newfrm
    p1, st, err = cv2.calcOpticalFlowPyrLK(GRold, GRnew, p0, None, **lk_params)
    if p1 is not None:
        Gnew = p1[st == 1]
        Gold = p0[st == 1]

    for i, (new, old) in enumerate(zip(Gnew, Gold)):
        a, b = new.ravel()
        c, d = old.ravel()

        listX.append(a-c)
        listY.append(b-d)
        mask = cv2.line(mask, (int(a), int(b)), (int(c), int(d)), color[i].tolist(), 2)
        frame = cv2.circle(frame, (int(a), int(b)), 5, color[i].tolist(), -1)
        img = cv2.add(frame, mask)
    list_X= np.array(listX)
    listX=reject_outliers(list_X)
    list_Y= np.array(listY)
    listY=reject_outliers(list_Y)
    LK_X=sum(listX) / len(listX)
    LK_Y=sum(listY) / len(listY)

    cv2.imwrite(newpath+'/LK/LK '+str(cnt)+'.jpg', img)
    TM=np.float32([[1, 0, int(LK_X)], [0, 1, int(LK_Y)]])
    return TM
'################'
def reject_outliers(data, m = 2.):
    d = np.abs(data - np.median(data))
    mdev = np.median(d)
    s = d/mdev if mdev else np.zeros(len(d))
    return data[s<m]

'################'
def warpPerspectivePadded(src, dst, transf):
    src_h, src_w = src.shape[:2]
    lin_homg_pts = np.array([[0, src_w, src_w, 0], [0, 0, src_h, src_h], [1, 1, 1, 1]])

    trans_lin_homg_pts = transf.dot(lin_homg_pts)

    minX = np.min(trans_lin_homg_pts[0, :])
    minY = np.min(trans_lin_homg_pts[1, :])
    maxX = np.max(trans_lin_homg_pts[0, :])
    maxY = np.max(trans_lin_homg_pts[1, :])

    dst_h, dst_w = dst.shape[:2]
    pad_h = int(np.round(np.maximum(dst_h, maxY) - np.minimum(0, minY)))
    pad_w = int(np.round(np.maximum(dst_w, maxX) - np.minimum(0, minX)))

    pad_sz = (pad_h, pad_w) + dst.shape[2:]
    dst_pad = np.zeros(pad_sz, dtype=np.uint8)

    anchorX = int(np.round(-minX)) if minX < 0 else 0
    anchorY = int(np.round(-minY)) if minY < 0 else 0

    transl_transf = np.eye(3)
    transl_transf[0, 2] += anchorX
    transl_transf[1, 2] += anchorY

    dst_pad[anchorY:anchorY + dst_h, anchorX:anchorX + dst_w] = dst

    return dst_pad


# Light Cyan Blue
lower1 = np.array([85, 100, 150])
upper1 = np.array([95, 255, 255])

# Light Blue
lower2 = np.array([100, 150, 200])
upper2 = np.array([120, 255, 255])

# Light Blue (another variation)
lower3 = np.array([120, 100, 150])
upper3 = np.array([130, 255, 255])

#blower = np.array([0, 0, 0])
#bupper = np.array([350, 66, 90])
blower = np.array([0, 0, 0])        # Lower limit: Black (no hue, saturation, or brightness)
bupper = np.array([360, 255, 50])
min_percent = 1  # Low percentile
max_percent = 95  # High percentile

ul, fl, mp =[np.zeros((map_h, map_w, 3)).astype(np.uint8) for _ in range(3)]
RF_r =[np.zeros((map_h, map_w, 3)).astype(np.uint8) for _ in range(1)]


#RFa = refpath(np.zeros(mp.shape), 1, thickness=1)  # .astype(np.uint8)
#mask_RFa = RFa != [0, 0, 0]
#RFb = refpath(np.zeros(mp.shape), 500, thickness=1)  # .astype(np.uint8)
#mask_RFb = RFb != [0, 0, 0]
#RF_ra = refpath(np.zeros(mp.shape), 1, color=(255, 0, 0), thickness=18).astype(np.uint8)
#mask_RF_ra = RF_ra != [0, 0, 0]
#RF_rb = refpath(np.zeros(mp.shape), 500, color=(255, 0, 0), thickness=28).astype(np.uint8)
#mask_RF_rb = RF_rb != [0, 0, 0]



static_text_img = pre_render_static_text(fontSc, thickn)
############################################################################################################################
for ki in File_name:
    Video_name=os.listdir(SourceF+'/'+str(ki))
    for kj in Video_name:
        start_t = time.time()
        cnt = -1  # 1150#1180
        flag=0
        v_name = SourceF+'/' + str(ki) + '/' + str(kj)
        tname = kj.split('.m')

        newpath = ResF + '/' + str(ki) + '/' + tname[0]
        if not os.path.exists(newpath):
            os.makedirs(newpath)
            os.makedirs(newpath + '/Nozzle')
            os.makedirs(newpath + '/Steps')
            os.makedirs(newpath + '/Mapping')
            os.makedirs(newpath + '/Mapping/Final')
            os.makedirs(newpath + '/Mapping/Undstr')
            os.makedirs(newpath + '/LK')



        Aw, L3, L4, L5 = [], [], [], []

        ## opening videocapture
        cap = cv2.VideoCapture(v_name)
        cap.set(cv2.CAP_PROP_POS_FRAMES, cnt)

        length = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        _, frame = cap.read()
        frame= image_resize(frame, resolution_w)
        ## some videowriter props
        sz = frame.shape

        fps = cap.get(cv2.CAP_PROP_FPS)

        fourcc = cv2.VideoWriter_fourcc('m', 'p', '4', 'v')

        ## open and set props
        vout = cv2.VideoWriter()
        vout.open(newpath+'/CCV analysis_' + tname[0] + '.mp4', fourcc, fps, (sz[1]*3, sz[0]+map_h), True)
        canvas_all = np.zeros((sz[0]+map_h, sz[1]*3, 3), dtype=np.uint8)
        #alpha = 0.8
        point_1=int(sz[0]*0.56)
        point_2=int(sz[0]*0.7)

        #length=630
        #cv2.imwrite('First frame ' + str(tname) + '.jpg', frame)
        circle = np.array([0])
        cir = np.array([0])
        newfrm, newink = [np.zeros((map_h, map_w)).astype(np.uint8) for _ in range(2)]
        dst_pad = np.zeros((map_h, int(sz[0]*3/2), 3), dtype=np.uint8)
        nr, nc = 720, 3300
        LK_TMold = np.zeros([2,3])
        LK_TMupd = np.zeros([2,3])
        cirr = []

        #fullp, Cafter = [np.random.randint(218, 228, size=(round(sz[1]), round(sz[0]*2 + 250), 3)).astype(np.uint8) for _ in range(2)]

        #Cbefore_ink, Cafter_ink = [np.zeros((round(sz[1]), round(sz[0]*2 + 250), 3)).astype(np.uint8) for _ in range(2)]

        #num_rows, num_cols = Cbefore_ink.shape[:2]

        #merge_skel=[]
        #merge=[]

        #lap_w=0
        #lap_h=0
        d , f = 0, 0
        #nextLK = 150
        while cnt < length-ftbp-1:
            if cnt<thdet:
                _, frame = cap.read()
                cnt += 1
            else:
                for i in range(ftbp):
                    _, frame = cap.read()
                    cnt += 1
            frame = image_resize(frame, resolution_w)
            original=frame.copy()
            cv2.putText(original, f'Frame #: {cnt}', (round(0.1 * resolution_w), round(0.2 * wdth_apprx)), cv2.FONT_HERSHEY_PLAIN, 2, (0,255,0), 2, cv2.LINE_AA)

            lo, hi = np.percentile(frame, (min_percent, max_percent))
            # Apply linear "stretch" - lo goes to 0, and hi goes to 1
            res_img = (frame.astype(float) - lo) / (hi - lo)
            # Multiply by 255, clamp range to [0, 255] and convert to uint8
            frame = np.maximum(np.minimum(res_img * 255, 255), 0).astype(np.uint8)

            if cnt > HTr[0] and cnt < HTr[1]:
                cir,im,mk = dscale(frame,blower,bupper)
                if cir is not None:
                    circle = cir
                    cirr.append(cir[2])
                    circle[2] = np.average(cirr)
                    mk = cv2.merge((mk, mk, mk))
                    HTres=np.concatenate((mk, im), axis=1)
                    cv2.imwrite(newpath+'/Nozzle/Nozzle detection example ' + str(cnt) + '_msk.jpg', HTres)
            if cnt==thdet-1:
                IAlineTh = np.round(((2 * circle[2]) / ref_wdth * 0.285) / (wdth_apprx / map_h)).astype(int)

        ###############################################################################################
        ################# ~ Width ~ #########################
            hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

            lower_mask = cv2.inRange(hsv, lower1, upper1)
            middle_mask = cv2.inRange(hsv, lower2, upper2)
            upper_mask = cv2.inRange(hsv, lower2, upper2)
            mask = lower_mask + middle_mask + upper_mask

            imask = mask > 0
            ink = np.zeros_like(frame, np.uint8)
            ink[imask] = frame[imask]
            ink_tmp=ink.copy()

            if circle.shape[0]>2:
                if cnt <= th:
                    avwidthc= Fwidth(mask, circle, cnt, newpath)
                    cv2.putText(ink, f'Width: {avwidthc} mm', (round(0.59 * resolution_w), round(0.5 * wdth_apprx)),
                                cv2.FONT_HERSHEY_SIMPLEX, fontSc, (0, 255, 0), thickn, cv2.LINE_AA)

                    if not math.isnan(avwidthc) and avwidthc<2:
                        Aw.append(avwidthc)

                else:
                    cv2.putText(ink, f'Av W:{round(np.average(Aw), 3)} mm', (round(0.3 * resolution_w), round(0.12 * wdth_apprx)),
                               cv2.FONT_HERSHEY_SIMPLEX, fontSc, (0, 255, 0), thickn, cv2.LINE_AA)

        #############################################################################################################################################
        ################# ~ Full print ~ #########################


            #############################################################################################################################################
            ###### ~ Compare ~ ########
            #CB = cv2.warpAffine(Cbefore_ink, translation_matrix, (num_cols, num_rows), borderValue=(0,0,0))
            #merge = CB#, Cafter_ink
            #merge_skel_= crtmap(merge)
            #if cnt == nextLK:
            oldfrm=newfrm
            #if dst_pad is None:
            oldink=newink
            #else:
            #    oldink=dst_pad
            newfrm = crtmap2(ink_tmp, 2)
            newink = ink_tmp.copy()
            newfrm[Nsp[0]:Nsp[1],Nep[0]:Nep[1]] = Zerofill
            #nextLK=cnt
            if oldfrm is not None and (cnt>thmp or cnt>obs_s):
                if cnt>=obs_s and flag==0:
                    dst_old=np.zeros_like(oldink)
                    dst_pad=np.zeros_like(oldink)
                    LK_TMupd=np.concatenate((LK_TM[:, :-1], np.zeros([2,1])), axis=1)
                    LK_TMold=np.concatenate((LK_TM[:, :-1], np.zeros([2,1])), axis=1)
                    flag=1
                if cnt<th or (cnt>=obs_s and cnt<obs_f):
                    if cnt<th: #and str(kj)=='000_Test1_w285h285_05ac_100pd_1400rpm.mp4':
                        LK_TM = mapping(oldfrm, newfrm, cnt)
                        LK.append(LK_TM)
                        lksm1 = sum(LK)[0][2]
                        lkr1=sum(LK)[0][0]
                        lksm2 = sum(LK)[0][2]*1.2#*0.50#((F1/F2))#*0.4887218045112782
                        lkr2=sum(LK)[0][0]*(obs_f-obs_s)/(th-thmp)
                        base_value_1 = lksm1 // lkr1
                        base_value_2 = lksm2 // lkr2
                        # Calculate the remainder
                        remainder_1 = lksm1 % lkr1
                        remainder_2 = lksm2 % lkr2
                        values1 = [base_value_1 + 1 if i < abs(int(remainder_1)) else base_value_1 for i in range(int(lkr1))]
                        values2 = [base_value_2 + 1 if i < abs(int(remainder_2)) else base_value_2 for i in range(int(lkr2))]
                    elif cnt>th:# and str(kj)=='000_Test1_w285h285_05ac_100pd_1400rpm.mp4':
                        #LM=random.choice(LK) #np.round(sum(LK) / len(LK))
                        #LK_TM = np.round(np.concatenate((LM[:, :-1], LM[:, -1].reshape(2, 1)*1.65), axis=1))
                        #print(tmp2)
                        LM=values2[tmp2]
                        tmp2+=1
                        tmp3+=1
                        LK_TM[0][2]=LM
                    elif cnt<th:
                        LM=values1[tmp1]
                        tmp1+=1
                        #LM=random.choice(LK) #np.round(sum(LK) / len(LK))
                        LK_TM[0][2]=LM#np.round(np.concatenate((LM[:, :-1], LM[:, -1].reshape(2, 1)), axis=1))
                        #if str(kj) == 'Test1_w285h285_05ac_100pd_1400rpm.mp4':
                        #    print('ok')
                    else:
                        #print(tmp2)
                        LM=values2[tmp2]
                        tmp2+=1
                        LK_TM[0][2]=LM
                        #LM = np.mean(LK, axis=0).round()
                        # LM=random.choice(LK) #np.round(sum(LK) / len(LK))
                        #LK_TM = np.round(np.concatenate((LM[:, :-1], LM[:, -1].reshape(2, 1)*1.62), axis=1)) #1.65

                    tm = LK_TM[:, -1] + LK_TMold[:, -1]
                    tm = tm.reshape(2, 1)
                    LK_TMupd = np.concatenate((LK_TM[:, :-1], tm), axis=1)
                    dst_pad = warpPerspectivePadded(oldink, newink, LK_TMupd)
                    h = np.min([dst_pad.shape[0], oldink.shape[0]])

                    if (LK_TM==LK_TMupd).all():
                        dst_pad[:h, :abs(int(LK_TMupd[0][2]))] = oldink[:h, :abs(int(LK_TM[0][2]))]
                    else:
                        #dst_pad[abs(int(LK_TMupd[1][2]))-abs(int(LK_TMold[1][2])):, :abs(int(LK_TMupd[0][2]))-abs(int(LK_TMold[0][2]))] = oldink[:, :abs(int(LK_TM[0][2]))]
                        #dst_pad[abs(int(LK_TMupd[1][2])) - abs(int(LK_TMold[1][2])):,abs(int(LK_TMupd[0][2])):] = oldink[:, abs(int(LK_TMold[0][2])):]

                        if int(LK_TMupd[1][2])==int(LK_TMold[1][2]):
                            dst_pad[:h,  abs(int(LK_TMold[0][2])):abs(int(LK_TMupd[0][2]))] = oldink[:h, :abs(int(LK_TM[0][2]))]
                            dst_pad[:h, :abs(int(LK_TMold[0][2]))]= dst_old[:h, :abs(int(LK_TMold[0][2]))]
                        else:
                            dst_pad[:h,  abs(int(LK_TMold[0][2])):abs(int(LK_TMupd[0][2]))] = oldink[:h, :abs(int(LK_TM[0][2]))]
                            dst_pad[:h, :abs(int(LK_TMold[0][2]))]= dst_old[:h, :abs(int(LK_TMold[0][2]))]
                    if cnt<th:
                        UL = cv2.warpAffine(dst_pad, tmul, (round(wdth_apprx*mapd), wdth_apprx),borderValue=(0, 0, 0))
                        ul = image_resize(UL, width=map_w)
                        cv2.imwrite(newpath + '/Mapping/Undstr/Mapping progress' + str(cnt) + '.jpg', ul)
                    elif cnt>obs_s and cnt<obs_f:
                        FL = cv2.warpAffine(dst_pad, tmfl, (round(wdth_apprx*mapd), wdth_apprx),borderValue=(0, 0, 0))
                        fl = image_resize(FL, width=map_w)
                        cv2.imwrite(newpath + '/Mapping/Final/Mapping progress' + str(cnt) + '.jpg', fl)
                    tm = LK_TM[:, -1] + LK_TMold[:, -1]
                    tm = tm.reshape(2, 1)
                    LK_TMold = np.concatenate((LK_TM[:, :-1], tm), axis=1)
                    dst_old = dst_pad.copy()

            #merge = dst_pad
            #merge_skel_= crtmap(dst_pad, )
            if cnt<th:
                mp = ul#merge#[0]#cv2.addWeighted(merge[0], alpha, merge[1], 1 - alpha, 0)
                msk = crtmap2(ul, 0)#merge_skel_#[0].copy()
                #if cnt==158:
                #    msk2=crtmap2(ul, 0)
                prv = mp#merge#[0].copy()
                prv_skl = msk#merge_skel_#[0].copy()
                mk=msk.copy()
            else:
                mp = fl#merge#[1]
                msk = crtmap2(fl, 1)#merge_skel_#[1].copy()
                mk=msk.copy()


            if cnt>=th:
                prv_ink = prv_skl != [0,0,0]
                msk[prv_ink]=prv_skl[prv_ink]

            mask_ink = mp != [0, 0, 0]
            # Color the pixels in the mask
            if cnt<th:
                RFa, prvpp = refpath(msk.copy(), cnt, prvpp, thickness=CDlineTh)  # .astype(np.uint8)
                mask_RFa = RFa != [0, 0, 0]
                RF_ra, prvpp = refpath(mp.copy(), cnt, prvpp, color=(255, 0, 0), thickness=IAlineTh)
                mask_RF_ra = RF_ra != [0, 0, 0]

                RF_rc = RF_ra.copy()
                msk[mask_RFa]=RFa.copy()[mask_RFa]
                RF_rc[mask_ink] = mp[mask_ink]
                RF_grey = cv2.cvtColor(RFa.astype(np.uint8), cv2.COLOR_RGB2GRAY)
                RF=RF_ra.copy()
            else:
                RFb, prvpp = refpath(msk.copy(), cnt, prvpp, thickness=CDlineTh)  # .astype(np.uint8)
                mask_RFb = RFb != [0, 0, 0]
                RF_rb, prvpp = refpath(mp.copy(), cnt, prvpp, color=(255, 0, 0), thickness=IAlineTh)
                mask_RF_rb = RF_rb != [0, 0, 0]

                RF_rc = RF_rb.copy()
                msk[mask_RFb]=RFb.copy()[mask_RFb]
                RF_rc[mask_ink] = mp[mask_ink]
                RF_grey = cv2.cvtColor(RFb.astype(np.uint8), cv2.COLOR_RGB2GRAY)
                RF=RF_rb.copy()

            if cnt>=th:
                RF_rc = cv2.addWeighted(prv, 0.6, RF_rc, 1 - 0.6, 0)
                msk = cv2.addWeighted(prv_skl, 0.2, msk, 1 - 0.2, 0)

            dd=RF_rc
            dd_sk=msk

            cv2.putText(dd_sk, 'Centerline accuracy (mm)', (round(0.05 * map_w), round(0.11 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,
                        (255, 255, 255), thickn, cv2.LINE_AA)

            if cnt > thmp:
                Q_grey = cv2.cvtColor(mk, cv2.COLOR_RGB2GRAY)
                bitwiseand = cv2.bitwise_and(RF_grey, Q_grey)
                _, bitwiseand = cv2.threshold(bitwiseand, 0, 255, cv2.THRESH_BINARY)

                Q_grey = cv2.cvtColor(mp.copy(), cv2.COLOR_RGB2GRAY)
                RF_r_grey = cv2.cvtColor(RF.astype(np.uint8), cv2.COLOR_RGB2GRAY)

                _, Q_grey = cv2.threshold(Q_grey, 0, 255, cv2.THRESH_BINARY)
                bitwiseand = cv2.bitwise_and(RF_r_grey, Q_grey)
                bitwiseor = cv2.bitwise_or(RF_r_grey, Q_grey)
                _, bitwiseand = cv2.threshold(bitwiseand, 0, 255, cv2.THRESH_BINARY)
                _, RF_r_bin = cv2.threshold(RF_r_grey, 0, 255, cv2.THRESH_BINARY)
                _, bitwiseor = cv2.threshold(bitwiseor, 0, 255, cv2.THRESH_BINARY)
                bitwiseor = bitwiseor - RF_r_bin
                #if cnt == length - 2:
                #    bw = bitwiseor[200:450, 0:1100]
                #    bw = image_resize(bw, width=1920)
                    # cv2.imwrite(newpath+'/2_Over-extruded area '+tname[0]+'.jpg', bw)

                total = RF_r_grey  # + Q_grey
                total_pixels = total[total > 0].shape[0]
                if total_pixels > 0:
                    matches = bitwiseand[bitwiseand > 0].shape[0]
                    matchesor = bitwiseor[bitwiseor > 0].shape[0]
                    percentage = round((100 * matches / total_pixels), 2)
                    percentageor = round((100 * matchesor / total_pixels), 2)

                else:
                    percentage = 0
                    percentageor = 0

                # print(percentage)
                # cv2.putText(dd, 'Area accuracy (%)', (10, 50), cv2.FONT_HERSHEY_PLAIN, fontSc,
                #            (255, 255, 255), thickn, cv2.LINE_AA)

                #error = Ferror2(RF_grey, Q_grey, circle, cnt)

                #cv2.imwrite('Overlap_' + str(i) + tname + '.jpg', bitwiseand)

                #total = RF_grey# + Q_grey
                #total_pixels = total[total > 0].shape[0]
                #if total_pixels>0:
                #    matches = bitwiseand[bitwiseand > 0].shape[0]
                #    percentage = round((100 * matches / total_pixels),2)
                #else:
                #    percentage = 0
                ##print(percentage)

                if cnt <th:
                    ULerror = Ferror(RF_grey, Q_grey, circle, ref_wdth)

                    er=ULerror.copy()
                    dd_sk = cv2.addWeighted(dd_sk, 1.0, static_text_img, 1.0, 0)
                    cv2.putText(dd_sk, f'{er} mm', (round(0.2 * map_w), round(0.89 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc, (0, 255, 0), thickn, cv2.LINE_AA)

                    cv2.putText(dd, 'Area accuracy (%)', (round(0.05 * map_w), round(0.11 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,(255, 255, 255), thickn, cv2.LINE_AA)

                    perc = percentage
                    percor = percentageor
                    #cv2.putText(dd, f'Undistrupted line: ', (round(0.05 * map_w), round(0.86*map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,(255, 255, 255), thickn, cv2.LINE_AA)
                    cv2.putText(dd, f'{perc} %', (round(0.2 * map_w), round(0.89*map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,(0, 255, 0), thickn, cv2.LINE_AA)
                    cv2.putText(dd, f'{percor} %', (round(0.36 * map_w), round(0.89*map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,(0, 0, 255), thickn, cv2.LINE_AA)
                elif cnt >th:
                    FLerror = Ferror(RF_grey, Q_grey, circle, ref_wdth)

                    dd_sk = draw_dynamic_texts(dd_sk, static_text_img, er, FLerror, fontSc, thickn)

                    cv2.putText(dd, 'Area accuracy (%)', (round(0.05 * map_w), round(0.11 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,(255, 255, 255), thickn, cv2.LINE_AA)


                    #cv2.putText(dd, f'Undistrupted line: ', (round(0.05 * map_w), round(0.86*map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,(255, 255, 255), thickn, cv2.LINE_AA)
                    cv2.putText(dd, f'{perc} %', (round(0.2 * map_w), round(0.89*map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,(0, 255, 0), thickn, cv2.LINE_AA)
                    cv2.putText(dd, f'{percor} %', (round(0.36 * map_w), round(0.89*map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc,(0, 0, 255), thickn, cv2.LINE_AA)

                    #cv2.putText(dd, 'Final path: ', (round(0.48*map_w), round(0.05 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc, (255, 255, 255), thickn, cv2.LINE_AA)
                    cv2.putText(dd, f'{percentage} %', (round(0.57*map_w), round(0.11 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc, (0, 255, 0), thickn, cv2.LINE_AA)
                    cv2.putText(dd, f'{percentageor} %', (round(0.73*map_w), round(0.11 * map_h)), cv2.FONT_HERSHEY_PLAIN, fontSc, (0, 0, 255), thickn, cv2.LINE_AA)



            '''
            RF_r_grey = cv2.cvtColor(RF_rc.astype(np.uint8), cv2.COLOR_RGB2GRAY)
            for i, cmp in enumerate(merge):
                Q_grey = cv2.cvtColor(cmp, cv2.COLOR_RGB2GRAY)
                _, Q_grey = cv2.threshold(Q_grey, 0, 255, cv2.THRESH_BINARY)
                bitwiseand = cv2.bitwise_and(RF_r_grey, Q_grey)
                bitwiseor = cv2.bitwise_or(RF_r_grey, Q_grey)
                _, bitwiseand = cv2.threshold(bitwiseand, 0, 255, cv2.THRESH_BINARY)
                _, RF_r_bin = cv2.threshold(RF_r_grey, 0, 255, cv2.THRESH_BINARY)
                _, bitwiseor = cv2.threshold(bitwiseor, 0, 255, cv2.THRESH_BINARY)
                bitwiseor = bitwiseor-RF_r_bin
                if cnt==length-2:
                    bw = bitwiseor[200:450, 0:1100]
                    bw = image_resize(bw, width=1920)
                    #cv2.imwrite(newpath+'/2_Over-extruded area '+tname[0]+'.jpg', bw)

                total = RF_r_grey# + Q_grey
                total_pixels = total[total > 0].shape[0]
                if total_pixels>0:
                    matches = bitwiseand[bitwiseand > 0].shape[0]
                    matchesor = bitwiseor[bitwiseor > 0].shape[0]
                    percentage = round((100 * matches / total_pixels), 2)
                    percentageor = round((100 * matchesor / total_pixels), 2)
                else:
                    percentage = 0
                    percentageor = 0

                # print(percentage)
                #cv2.putText(dd, 'Area accuracy (%)', (10, 50), cv2.FONT_HERSHEY_PLAIN, fontSc,
                #            (255, 255, 255), thickn, cv2.LINE_AA)
                if i==0 and cnt <421:
                    perc=percentage
                    percor=percentageor
                    #cv2.putText(dd, 'Undistrupted line: ', (10, 400), cv2.FONT_HERSHEY_PLAIN, fontSc,
                    #            (255, 255, 255), thickn, cv2.LINE_AA)
                    #cv2.putText(dd, str(perc) +'%', (640, 400), cv2.FONT_HERSHEY_PLAIN, fontSc,
                    #            (0, 255, 0), thickn, cv2.LINE_AA)
                    #cv2.putText(dd, str(percor) + '%', (890, 400), cv2.FONT_HERSHEY_PLAIN, fontSc,
                    #            (0, 0, 255), thickn, cv2.LINE_AA)
                #elif i == 1 and cnt > 421:
                #    cv2.putText(dd, 'Undistrupted line: ', (10, 400), cv2.FONT_HERSHEY_PLAIN, fontSc,
                #                (255, 255, 255), thickn, cv2.LINE_AA)
                #    cv2.putText(dd, str(perc) +'%', (640, 400), cv2.FONT_HERSHEY_PLAIN, fontSc,
                #                (0, 255, 0), thickn, cv2.LINE_AA)
                #    cv2.putText(dd, str(percor) + '%', (890, 400), cv2.FONT_HERSHEY_PLAIN, fontSc,
                #                (0, 0, 255), thickn, cv2.LINE_AA)
                #    cv2.putText(dd, 'Final path: ', (1070, 50), cv2.FONT_HERSHEY_PLAIN, fontSc, (255, 255, 255), thickn, cv2.LINE_AA)
                #    cv2.putText(dd, str(percentage) +'%', (1440, 50), cv2.FONT_HERSHEY_PLAIN, fontSc, (0, 255, 0), thickn, cv2.LINE_AA)
                #    cv2.putText(dd, str(percentageor) +'%', (1690, 50), cv2.FONT_HERSHEY_PLAIN, fontSc, (0, 0, 255), thickn, cv2.LINE_AA)
            '''
            canvas_all[:sz[0], :sz[1]] = original
            canvas_all[:sz[0], sz[1]:sz[1]*2] = frame#dst
            canvas_all[:sz[0], sz[1]*2:] = ink#dst#sb#mm
            canvas_all[sz[0]:, :int(sz[1]*3/2)] = dd
            canvas_all[sz[0]:, int(sz[1]*3/2):] = dd_sk
            #if cnt>=156:
            #    print('ok')

            if cnt==110:
                cv2.imwrite(newpath+'/Nozzle/Frame_'+str(cnt)+'_' + tname[0] + '.jpg', original)
            elif cnt==420 or cnt==length-2:
                cv2.imwrite(newpath+'/0_Line_'+str(cnt)+'_ovl_' + tname[0] + '.jpg', dd)
                cv2.imwrite(newpath+'/1_Line_'+str(cnt)+'_avr_' + tname[0] + '.jpg', dd_sk)

                #white
                dd[np.where((dd==[0,0,0]).all(axis=2))] = [255,255,255]
                dd_sk[np.where((dd_sk==[0,0,0]).all(axis=2))] = [255,255,255]
                cv2.imwrite(newpath+'/0_Line_'+str(cnt)+'_ovl_W_' + tname[0] + '.jpg', dd)
                cv2.imwrite(newpath+'/1_Line_'+str(cnt)+'_avr_W_' + tname[0] + '.jpg', dd_sk)
            elif cnt==156:
                cv2.imwrite(newpath+'/Steps/Original_'+str(cnt)+'_' + tname[0] + '.jpg', original)
                cv2.imwrite(newpath+'/Steps/Ink_'+str(cnt)+'_' + tname[0] + '.jpg', ink)
                #cv2.imwrite(newpath+'/Steps/SIoI_'+str(cnt)+'_' + tname[0] + '.jpg', dst)
                cv2.imwrite(newpath+'/Steps/PerFr_'+str(cnt)+'_' + tname[0] + '.jpg', frame)
                cv2.imwrite(newpath+'/Steps/0_Overl_'+str(cnt)+'_' + tname[0] + '.jpg', dd)
                cv2.imwrite(newpath+'/Steps/1_AvE_'+str(cnt)+'_' + tname[0] + '.jpg', dd_sk)

            #os.system('cls')
            #end_t = time.time()
            #print('Loop time:', round(end_t - start_t, 5))
            #start_t = time.time()

            print(str(cnt) + str('/') + str(length - 1) + ' frame of file: ' + kj)
            vout.write(canvas_all)
            if cnt>=length-ftbp-1:
                data2 = pd.DataFrame({"Print": [kj], "Average width": [round(np.average(Aw), 3)], "U Area accuracy ": [perc], "U Excess area": [percor], "F Area accuracy ": [percentage], "F Excess area": [percentageor], "U Centerline acc": [er], "F Centerline acc": [FLerror]})
                dfcmp=pd.concat([df, data2], ignore_index=True)
                dfcmp.to_excel(excel, sheet_name='Print quality results', index=False)
                df=dfcmp
        ul, fl, mp = [np.zeros((map_h, int(resolution_w * 3 / 2), 3)).astype(np.uint8) for _ in range(3)]
        RF_r = [np.zeros((map_h, int(resolution_w * 3 / 2), 3)).astype(np.uint8) for _ in range(1)]
        end_t = time.time()
        print(str(cnt) + str('/') + str(length - 1) + ' frame of file: ' + kj)
        print('Loop time:', round(end_t - start_t, 5))
        tmp1=0
        tmp2=0
        vout.release()
        cap.release()

# Step 1: Rearrange the first 9 rows
new_order = [6, 7, 8, 3, 4, 5, 0, 1, 2]  # Rearranging the first 9 rows
rearranged_part = df.iloc[new_order].reset_index(drop=True)  # Rearrange first 9 rows
rest_part = df.iloc[9:].reset_index(drop=True)  # Keep the rest unchanged

# Combine rearranged first 9 rows with the rest of the DataFrame
rearranged_rows = pd.concat([rearranged_part, rest_part], ignore_index=True)

# Step 2: Insert averages after every three rows, for the entire DataFrame (rearranged_rows)
new_rows = []

# Process the entire rearranged DataFrame
for i in range(0, len(rearranged_rows), 3):  # Iterate over in steps of 3
    # Add current group of 3 rows
    new_rows.append(rearranged_rows.iloc[i:i+3])

    # Calculate the average if there are 3 rows in this chunk
    if i + 3 <= len(rearranged_rows):
        avg_row = rearranged_rows.iloc[i:i+3, 1:].mean(numeric_only=True)
        avg_row_df = pd.DataFrame([['Average'] + avg_row.tolist()], columns=rearranged_rows.columns)
        new_rows.append(avg_row_df)

# Concatenate all the rows including the averages
df_with_avgs = pd.concat(new_rows, ignore_index=True)

# Display the modified DataFrame with average rows inserted
print("\nModified DataFrame with Averages:")
print(df_with_avgs)

df_with_avgs.to_excel(excel, sheet_name='Print quality results', index=False)

# Define the light blue fill
light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

# Create an Excel writer using openpyxl
output_file = excel
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_with_avgs.to_excel(writer, index=False, sheet_name='Print quality results')

    # Get the workbook and the worksheet
    workbook = writer.book
    worksheet = writer.sheets['Print quality results']

    # Apply the fill color to the average rows
    for row in range(1, len(df_with_avgs) + 1):  # Starting from 1 to skip header
        if df_with_avgs.iloc[row - 1, 0] == 'Average':  # Check if the first column is 'Average'
            for col in range(1, len(df_with_avgs.columns) + 1):  # Fill columns B, C, D
                worksheet.cell(row=row + 1, column=col).fill = light_blue_fill  # +1 for header row

print(f"\nDataFrame successfully written to '{output_file}' with averages highlighted.")